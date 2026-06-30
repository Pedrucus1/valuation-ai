#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Replica la metodologia PURA del perito leyendo sus factores reales del OPI Constr,
y la compara con una version AUTOMATABLE (solo factores por formula, sin zona/ubic/acabados)."""
import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

def col(ws,letter,row): return ws[f'{letter}{row}'].value
def n(v):
    try: return float(v)
    except: return None

def find_row(ws, text, colrange=('A','AL','AI','C')):
    for r in range(1, ws.max_row+1):
        for cl in colrange:
            v=ws[f'{cl}{r}'].value
            if isinstance(v,str) and text in v.upper():
                return r
    return None

def analiza(path, label):
    wb=openpyxl.load_workbook(path, data_only=True)
    ws=wb['OPI Constr']
    # sujeto
    m2T=n(col(ws,'I',52)); m2C=n(col(ws,'I',54))
    rH=find_row(ws,'TABLA DE HOMOLOG')
    rC=find_row(ws,'TABLA DE CALIFIC')
    rFinal=find_row(ws,'RESULTADO POR EL M')
    if not(rH and rC):
        print(f'{label}: no encontre tablas'); return None
    # comps homologacion: filas rH+2 .. hasta PROMEDIO (5 max)
    comps=[]
    for r in range(rH+2, rH+8):
        s=n(col(ws,'C',r)); c=n(col(ws,'F',r)); p=n(col(ws,'I',r)); pm=n(col(ws,'AD',r))
        if s and c and p: comps.append(dict(s=s,c=c,p=p,pm2T=pm,row_h=r))
    # calificacion: alinear por indice (mismas filas relativas)
    for i,cp in enumerate(comps):
        r=rC+2+i
        cp['zona']=n(col(ws,'L',r)); cp['ubic']=n(col(ws,'N',r)); cp['cusF']=n(col(ws,'Q',r))
        cp['supF']=n(col(ws,'T',r)); cp['edadF']=n(col(ws,'X',r)); cp['consF']=n(col(ws,'Y',r))
        cp['acabF']=n(col(ws,'AB',r)); cp['otroF']=n(col(ws,'AE',r)); cp['AH']=n(col(ws,'AH',r))
        cp['AL_perito']=n(col(ws,'AL',r))
    valorPerito = n(col(ws,'AJ',rFinal)) if rFinal else None
    return dict(label=label,m2T=m2T,m2C=m2C,comps=comps,valorPerito=valorPerito)

def replica(d):
    m2C=d['m2C']; m2T=d['m2T']; cus_s=m2C/m2T
    pm2T = d['comps'][0]['pm2T'] or 0
    pur=[]; auto=[]
    for cp in d['comps']:
        s,c,p=cp['s'],cp['c'],cp['p']; cus_c=c/s
        pu_raw=p/c
        pu_homol=pu_raw + pm2T*(1/cus_s - 1/cus_c)
        # PURA: con factores reales del perito (paso2 completo)
        AH = cp['AH']
        pur.append(pu_homol*AH)
        # AUTOMATABLE: solo formulas (CUS^1/6, sup-terreno^1/6, edad real del perito), sin zona/ubic/acab/otro/conserv
        fCUS=(cus_s/cus_c)**(1/6); fSup=(m2T/s)**(1/6)
        fEdad=cp['edadF'] or 1.0
        auto.append(pu_homol*fCUS*fSup*fEdad)
    pu_pur=sum(pur)/len(pur); pu_auto=sum(auto)/len(auto)
    return pu_pur*m2C, pu_auto*m2C, pm2T

archivos=[('opi_26-5-16.xlsx','26-5-16'),('opi_613.xlsx','26-6-13'),('opi_2512-02.xlsx','25-12-02'),
          ('opi_257-14.xlsx','25-7-14'),('opi_259-01.xlsx','25-9-01'),('opi_259-02.xlsx','25-9-02'),
          ('opi_2511-07.xlsx','25-11-07')]
import os
print(f"{'OPI':<12}{'m2T':>5}{'m2C':>5}{'CUS':>6}{'pm2T':>8}{'PERITO':>10}{'REPLICA':>10}{'err%':>7}{'AUTO':>10}{'errAuto':>8}")
print('-'*88)
for f,lbl in archivos:
    if not os.path.exists(f): continue
    d=analiza(f,lbl)
    if not d or not d['comps'] or not d['valorPerito']:
        print(f'{lbl}: datos incompletos'); continue
    vp,va,pm=replica(d)
    per=d['valorPerito']
    cus=d['m2C']/d['m2T']
    print(f"{lbl:<12}{d['m2T']:>5.0f}{d['m2C']:>5.0f}{cus:>6.2f}{pm:>8,.0f}{per:>10,.0f}{vp:>10,.0f}"
          f"{(vp/per-1)*100:>+7.1f}{va:>10,.0f}{(va/per-1)*100:>+8.1f}")
