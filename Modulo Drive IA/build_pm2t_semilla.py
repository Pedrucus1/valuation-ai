#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Construye pm2t_semilla.json: valor de terreno ($/m2) por colonia,
sembrado desde los comparables de TERRENO que el perito uso en sus 768 avaluos (cerebro_datos.json).
Tambien incorpora el AC108 (valor terreno) de las hojas opi_*.xlsx descargadas si existen.
Pondera por antiguedad: OPIs mas recientes pesan mas (decae cada 3 meses, ver peso())."""
import json, re, sys, unicodedata, statistics as st, glob, os, datetime
sys.stdout.reconfigure(encoding='utf-8')

def folio_edad_meses(folio):
    """Meses desde hoy hasta la fecha codificada en el folio OPI-YY-M-D-... None si no matchea."""
    m = re.match(r'OPI-(\d{2})-(\d{1,2})-', str(folio or ''))
    if not m: return None
    anio, mes = 2000 + int(m.group(1)), int(m.group(2))
    hoy = datetime.date.today()
    return max(0, (hoy.year - anio) * 12 + (hoy.month - mes))

def peso(edad_meses):
    """1.0 los primeros 3 meses, luego decae por trimestre (0.5, 0.33, 0.25, ...)."""
    if edad_meses is None: return 1.0
    return 1.0 / (1 + edad_meses // 3)

def mediana_ponderada(pares):
    """pares = [(valor, peso), ...] -> valor en el punto donde el peso acumulado cruza 50%."""
    pares = sorted(pares, key=lambda p: p[0])
    total = sum(w for _, w in pares)
    acc = 0.0
    for v, w in pares:
        acc += w
        if acc >= total / 2:
            return v
    return pares[-1][0]

def norm(s):
    s=unicodedata.normalize('NFD',str(s or '')).encode('ascii','ignore').decode().lower()
    s=re.sub(r'\b(col\.|colonia|fracc\.|fraccionamiento|residencial|secc?\.?|coto|privada|conjunto)\b',' ',s)
    return re.sub(r'[^a-z0-9 ]',' ',s).replace('  ',' ').strip()
def nm(s):
    m=norm(s); m=re.sub(r'tlajomulco de zuniga','tlajomulco',m); m=re.sub(r'^san pedro.*','tlaquepaque',m); return m.strip()
def num(v):
    try: return float(re.sub(r'[^0-9.\-]','',str(v))) if v not in(None,'') else 0.0
    except: return 0.0

cer=json.load(open('cerebro_datos.json',encoding='utf-8'))
# acumular pm2T por (muni|colonia) desde comps de terreno del perito, con peso por antiguedad de la OPI
acc={}
for x in cer:
    muni=nm(x.get('municipio','')); col=norm(x.get('sujetoColonia',''))
    if not muni or len(col)<4: continue
    w=peso(folio_edad_meses(x.get('folio')))
    terr=[c for c in x.get('comparables',[]) if num(c.get('construccion'))==0 and num(c.get('terreno'))>0 and num(c.get('precio'))>0]
    vals=[num(c['precio'])/num(c['terreno']) for c in terr]
    vals=[v for v in vals if 1500<=v<=30000]  # rango plausible de suelo
    if not vals: continue
    key=f"{muni}|{col}"
    acc.setdefault(key,[]).extend((v,w) for v in vals)

semilla={}
for key,pares in acc.items():
    if len(pares)<1: continue
    semilla[key]={'pm2T':round(mediana_ponderada(pares)), 'n':len(pares), 'fuente':'perito_terreno_comps'}

# incorporar AC108 de hojas descargadas (valor terreno homologado del perito, mas confiable)
try:
    import openpyxl
    for f in glob.glob('opi_*.xlsx'):
        try:
            wb=openpyxl.load_workbook(f,data_only=True)
            if 'OPI Constr' not in wb.sheetnames: continue
            ws=wb['OPI Constr']
            ac=num(ws['AC108'].value); col=norm(ws['AG74'].value); muni=nm(ws['AB72'].value)
            if ac and len(col)>=4 and muni and 1500<=ac<=30000:
                key=f"{muni}|{col}"
                semilla[key]={'pm2T':round(ac),'n':semilla.get(key,{}).get('n',0)+1,'fuente':'perito_AC108'}  # AC108 pisa (mas confiable)
        except Exception: pass
except ImportError: pass

out={'_meta':{'fecha':'2026-06-30','total_colonias':len(semilla),'fuente':'cerebro terreno comps + AC108 hojas perito'},
     'zonas':semilla}
json.dump(out, open('pm2t_semilla.json','w',encoding='utf-8'), ensure_ascii=False, indent=1)
print(f"pm2t_semilla.json escrito: {len(semilla)} colonias con valor de terreno")
print("ejemplos:")
for k,v in list(semilla.items())[:10]:
    print(f"  {k:<40} ${v['pm2T']:>7,} (n={v['n']}, {v['fuente']})")
