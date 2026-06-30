#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Descompone los factores de calificacion del perito: separa la parte AUTOMATABLE
(CUS,superficie,edad) de la parte MANUAL (zona,ubicacion,conservacion,acabados,otro)
para ver si la manual es derivable (vs conservacion del sujeto) o una constante."""
import openpyxl, sys, glob, os
sys.stdout.reconfigure(encoding='utf-8')
def n(v):
    try: return float(v)
    except: return None
def find_row(ws,t):
    for r in range(1,min(ws.max_row,260)+1):
        for cl in ('A','AI'):
            v=ws[f'{cl}{r}'].value
            if isinstance(v,str) and t in v.upper(): return r

print(f"{'archivo':<13}{'CUS':>5}{'consSuj':<14}{'manualProm':>11}{'zona':>6}{'ubic':>6}{'cons':>6}{'acab':>6}{'otro':>6}")
data=[]
for f in sorted(glob.glob('opi_*.xlsx')):
    wb=openpyxl.load_workbook(f,data_only=True)
    if 'OPI Constr' not in wb.sheetnames: continue
    ws=wb['OPI Constr']
    m2T=n(ws['I52'].value); m2C=n(ws['I54'].value)
    if not m2T or not m2C: continue
    consSuj=ws['W86'].value or ws['X48'].value
    rC=find_row(ws,'TABLA DE CALIFIC')
    if not rC: continue
    L=N=Y=AB=AE=man=cnt=0
    for i in range(5):
        r=rC+2+i
        l=n(ws['L'+str(r)].value);u=n(ws['N'+str(r)].value);y=n(ws['Y'+str(r)].value)
        a=n(ws['AB'+str(r)].value);o=n(ws['AE'+str(r)].value)
        if None in (l,u,y,a,o): continue
        L+=l;N+=u;Y+=y;AB+=a;AE+=o; man+= l*u*y*a*o; cnt+=1
    if not cnt: continue
    cus=m2C/m2T
    data.append((os.path.basename(f).replace('.xlsx',''),cus,str(consSuj),man/cnt,L/cnt,N/cnt,Y/cnt,AB/cnt,AE/cnt))

for d in sorted(data,key=lambda x:x[1]):
    print(f"{d[0]:<13}{d[1]:>5.2f}{d[2][:13]:<14}{d[3]:>11.3f}{d[4]:>6.2f}{d[5]:>6.2f}{d[6]:>6.2f}{d[7]:>6.2f}{d[8]:>6.2f}")

import statistics as st
mans=[d[3] for d in data]
print(f"\nFactor MANUAL promedio: {st.mean(mans):.3f}  min {min(mans):.3f}  max {max(mans):.3f}  stdev {st.pstdev(mans):.3f}")
# agrupar por conservacion del sujeto
print('\nFactor manual por estado de conservacion del sujeto:')
byc={}
for d in data: byc.setdefault(d[2],[]).append(d[3])
for k,v in sorted(byc.items()): print(f"  {k[:22]:<24} n={len(v)} manual_prom={st.mean(v):.3f}")
