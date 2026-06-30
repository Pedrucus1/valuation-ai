#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Calibracion metodologia perito + seed de pm2T por zona, sobre todos los opi_*.xlsx.
pm2T se toma de AC108 (valor resultante del terreno/m2) de OPI Constr (regla del usuario)."""
import openpyxl, sys, glob, os
sys.stdout.reconfigure(encoding='utf-8')

def n(v):
    try: return float(v)
    except: return None

def find_row(ws,t):
    for r in range(1,min(ws.max_row,260)+1):
        for cl in ('A','AI'):
            v=ws[f'{cl}{r}'].value
            if isinstance(v,str) and t in v.upper(): return r

def analiza(path):
    wb=openpyxl.load_workbook(path,data_only=True)
    if 'OPI Constr' not in wb.sheetnames: return None
    ws=wb['OPI Constr']
    m2T=n(ws['I52'].value); m2C=n(ws['I54'].value)
    if not m2T or not m2C: return None
    colonia=ws['AG74'].value; muni=ws['AB72'].value; tipo=ws['E8'].value
    pm2T_AC=n(ws['AC108'].value)
    rH=find_row(ws,'TABLA DE HOMOLOG'); rC=find_row(ws,'TABLA DE CALIFIC'); rF=find_row(ws,'RESULTADO POR EL M')
    if not(rH and rC and rF): return None
    comps=[]
    for i in range(5):
        r=rH+2+i; s=n(ws['C'+str(r)].value); c=n(ws['F'+str(r)].value); p=n(ws['I'+str(r)].value); pm=n(ws['AD'+str(r)].value)
        if s and c and p:
            rc=rC+2+i
            comps.append(dict(s=s,c=c,p=p,pm=pm,AH=n(ws['AH'+str(rc)].value),edadF=n(ws['X'+str(rc)].value)))
    per=n(ws['AJ'+str(rF)].value)
    cons=ws['W86'].value or ws['X48'].value
    return dict(m2T=m2T,m2C=m2C,colonia=colonia,muni=muni,tipo=tipo,pm2T=pm2T_AC,comps=comps,per=per,cons=cons)

# calibracion del factor manual omitido (zona/ubic/conserv/acab/otro) segun conservacion del sujeto
CAL={'BUENO':0.95,'REGULAR':0.89,'REMODELADO':0.90,'REMODELADA':0.90}
def cal_factor(cons):
    s=(cons or '').upper()
    for k,v in CAL.items():
        if k in s: return v
    return 0.92

def replica(d):
    m2C,m2T=d['m2C'],d['m2T']; cus_s=m2C/m2T; pm=d['pm2T'] or 0
    pur=[]; auto=[]
    for cp in d['comps']:
        s,c,p=cp['s'],cp['c'],cp['p']; cus_c=c/s; pu_raw=p/c
        pu_h=pu_raw + pm*(1/cus_s - 1/cus_c)
        if cp['AH']: pur.append(pu_h*cp['AH'])
        fe=cp['edadF'] or 1.0
        auto.append(pu_h*(cus_s/cus_c)**(1/6)*(m2T/s)**(1/6)*fe)
    vp=sum(pur)/len(pur)*m2C if pur else None
    va=sum(auto)/len(auto)*m2C
    vcal=va*cal_factor(d.get('cons'))
    return vp,va,vcal

rows=[]
for f in sorted(glob.glob('opi_*.xlsx')):
    d=analiza(f)
    if not d or not d['comps'] or not d['per']: continue
    vp,va,vcal=replica(d)
    if not vp: continue
    cus=d['m2C']/d['m2T']
    rows.append(dict(f=os.path.basename(f),muni=(d['muni'] or '')[:14],col=(d['colonia'] or '')[:20],
                     cus=cus,pm2T=d['pm2T'] or 0,per=d['per'],vp=vp,va=va,vcal=vcal,
                     ep=(vp/d['per']-1)*100, ea=(va/d['per']-1)*100, ec=(vcal/d['per']-1)*100))

print(f'\n=== CALIBRACION ({len(rows)} OPIs casa) — err vs perito ===')
print(f"{'archivo':<14}{'CUS':>5}{'pm2T':>8}{'PERITO':>10}{'PURA':>8}{'AUTO':>8}{'AUTO+cal':>9}")
for r in sorted(rows,key=lambda r:r['cus']):
    print(f"{r['f'].replace('.xlsx',''):<14}{r['cus']:>5.2f}{r['pm2T']:>8,.0f}{r['per']:>10,.0f}{r['ep']:>+7.1f}{r['ea']:>+8.1f}{r['ec']:>+9.1f}")
import statistics as st
puras=[abs(r['ep']) for r in rows]; autos=[abs(r['ea']) for r in rows]; cals=[abs(r['ec']) for r in rows]
print(f"\nError abs PURA:     prom {st.mean(puras):.2f}%  max {max(puras):.1f}%")
print(f"Error abs AUTO:     prom {st.mean(autos):.2f}%  mediana {st.median(autos):.1f}%  max {max(autos):.1f}%")
print(f"Error abs AUTO+cal: prom {st.mean(cals):.2f}%  mediana {st.median(cals):.1f}%  max {max(cals):.1f}%")
print(f"AUTO+cal dentro de +-10%: {sum(1 for a in cals if a<=10)}/{len(cals)}  | +-15%: {sum(1 for a in cals if a<=15)}/{len(cals)}")

print('\n=== SEED pm2T por zona (de AC108) ===')
print(f"{'municipio':<15}{'colonia':<22}{'pm2T':>9}")
for r in sorted(rows,key=lambda r:(r['muni'],-r['pm2T'])):
    print(f"{r['muni']:<15}{r['col']:<22}{r['pm2T']:>9,.0f}")
