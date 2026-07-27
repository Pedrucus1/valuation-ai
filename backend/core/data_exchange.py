"""Data Exchange (#133) — lógica pura: plantilla, parseo y validación por tipo.

La inmobiliaria llena NUESTRA plantilla (columnas fijas), así el parseo es
determinista: no hace falta IA para mapear encabezados. Una fila válida se
escribe luego a dos lugares (CRM `propiedades_inmobiliaria` + pool
`mercado_props`); eso lo hace el router. Aquí solo: generar plantilla, leer el
archivo subido, normalizar cada fila y validar los obligatorios SEGÚN el tipo.
"""
from __future__ import annotations
import csv
import hashlib
import io
import re
from datetime import date

# Columna canónica (clave interna) → etiqueta que ve la inmobiliaria en la plantilla.
# El ORDEN es el de la plantilla; Tipo va primero (define qué se exige).
COLUMNAS = [
    ("tipo",            "Tipo de propiedad"),
    ("direccion",       "Dirección"),
    ("coto_edificio",   "Coto / Edificio"),
    ("colonia",         "Colonia"),
    ("municipio",       "Municipio"),
    ("precio",          "Precio de salida (MXN)"),
    ("anio",            "Año de construcción"),
    ("edad",            "Edad (años)"),
    ("m2_construccion", "m² construcción"),
    ("m2_terreno",      "m² terreno"),
    ("recamaras",       "Recámaras"),
    ("banos",           "Baños"),
    ("medios_banos",    "Medios baños"),
    ("estacionamientos","Estacionamientos"),
    ("niveles",         "Niveles"),
    ("piso",            "Piso"),
    ("conservacion",    "Estado de conservación"),
    ("grado_remodelacion", "Grado de remodelación (opcional)"),
    ("anio_remodelacion",  "Año de remodelación (opcional)"),
    ("amenidades",      "Amenidades"),
    ("descripcion",     "Descripción"),
    ("link",            "Link de origen (opcional)"),
]
ETIQUETA = dict(COLUMNAS)                       # clave → etiqueta
CLAVE_POR_ETIQUETA = {v.lower(): k for k, v in COLUMNAS}

TIPOS_CANON = {
    "casa": "casa", "departamento": "departamento", "depto": "departamento",
    "terreno": "terreno", "lote": "terreno",
    "local": "local", "local comercial": "local", "oficina": "oficina",
    "bodega": "bodega",
}

# Obligatorios por tipo (claves canónicas). "_edad" = requiere año O edad.
NUMERICOS = {"precio", "anio", "edad", "m2_construccion", "m2_terreno",
             "recamaras", "banos", "medios_banos", "estacionamientos", "niveles", "piso",
             "anio_remodelacion"}
_BASE = {"tipo", "direccion", "colonia", "municipio", "precio"}
REQUERIDOS_POR_TIPO = {
    "casa":        _BASE | {"_edad", "m2_construccion", "m2_terreno", "recamaras",
                            "banos", "medios_banos", "estacionamientos", "niveles", "conservacion"},
    "departamento":_BASE | {"_edad", "m2_construccion", "m2_terreno", "recamaras",
                            "banos", "medios_banos", "estacionamientos", "niveles", "conservacion"},
    "terreno":     _BASE | {"m2_terreno"},
    "local":       _BASE | {"_edad", "m2_construccion", "m2_terreno", "estacionamientos", "conservacion"},
    "oficina":     _BASE | {"_edad", "m2_construccion", "m2_terreno", "estacionamientos", "conservacion"},
    "bodega":      _BASE | {"_edad", "m2_construccion", "m2_terreno", "estacionamientos", "conservacion"},
}

EJEMPLO = {
    "tipo": "Casa", "direccion": "Av. Patria 1200", "coto_edificio": "Coto Las Palmas",
    "colonia": "Jardines de San Ignacio",
    "municipio": "Zapopan", "precio": "3500000", "anio": "2015", "edad": "",
    "m2_construccion": "180", "m2_terreno": "200", "recamaras": "3", "banos": "2",
    "medios_banos": "1", "estacionamientos": "2", "niveles": "2", "piso": "",
    "conservacion": "Bueno", "amenidades": "Alberca, Seguridad 24h, Panel solar",
    "descripcion": "Casa en excelente ubicación",
}


def _num(v):
    """'3,500,000' / '$3500000' / '180 m²' → float; None si vacío/no numérico."""
    if v is None:
        return None
    s = re.sub(r"[^\d.]", "", str(v).replace(",", ""))
    if not s or s == ".":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def normalizar_tipo(v) -> str | None:
    return TIPOS_CANON.get(str(v or "").strip().lower())


def normalizar_fila(raw: dict) -> dict:
    """raw: clave_canónica → valor crudo. Devuelve fila normalizada con `anio`
    derivado de edad si hace falta. No valida (eso es validar_fila)."""
    out = {}
    out["tipo"] = normalizar_tipo(raw.get("tipo"))
    for k in ("direccion", "coto_edificio", "colonia", "municipio", "conservacion",
              "grado_remodelacion", "amenidades", "descripcion", "link"):
        val = str(raw.get(k) or "").strip()
        out[k] = val or None
    for k in NUMERICOS:
        out[k] = _num(raw.get(k))
    # Edad → año (canónico es el año). Año explícito válido manda.
    anio, edad, ahora = out.get("anio"), out.get("edad"), date.today().year
    if not (anio and 1800 <= anio <= ahora + 2):
        out["anio"] = int(ahora - edad) if edad and 0 <= edad <= 200 else None
    else:
        out["anio"] = int(anio)
    return out


def validar_fila(fila: dict) -> list[str]:
    """Devuelve las etiquetas de campos obligatorios FALTANTES según el tipo.
    Lista vacía = fila válida. Tipo desconocido = error propio."""
    tipo = fila.get("tipo")
    if tipo not in REQUERIDOS_POR_TIPO:
        return [f"Tipo de propiedad inválido ('{fila.get('tipo')}')"]
    faltan = []
    for req in REQUERIDOS_POR_TIPO[tipo]:
        if req == "_edad":
            if not fila.get("anio"):
                faltan.append("Año de construcción o Edad")
        elif fila.get(req) in (None, ""):
            faltan.append(ETIQUETA[req])
    return faltan


def id_unico_data_exchange(inmobiliaria_id: str, direccion: str) -> str:
    base = f"{inmobiliaria_id}|{(direccion or '').strip().lower()}"
    return hashlib.md5(base.encode("utf-8")).hexdigest()


def clave_direccion(direccion: str) -> str:
    """Clave normalizada de dirección para deduplicar (colapsa espacios/caso)."""
    return " ".join(str(direccion or "").lower().split())


_TIPO_DISPLAY = {"casa": "Casa", "departamento": "Departamento", "terreno": "Terreno",
                  "local": "Local", "oficina": "Oficina", "bodega": "Bodega"}


def fila_a_doc_crm(f: dict, user_id: str, *, origen: str, ahora: str) -> dict:
    """Fila normalizada → doc de `propiedades_inmobiliaria` (CRM de la inmobiliaria)."""
    anio_actual = date.today().year
    return {
        "user_id": user_id, "origen": origen,
        "direccion": f["direccion"], "tipo": _TIPO_DISPLAY.get(f["tipo"], f["tipo"]),
        "coto_edificio": f.get("coto_edificio"), "piso": f.get("piso"),
        "amenidades": f.get("amenidades"),
        "colonia": f["colonia"], "municipio": f["municipio"],
        "precio_oferta": f["precio"],
        "m2_construccion": f.get("m2_construccion"), "m2_terreno": f.get("m2_terreno"),
        "recamaras": f.get("recamaras"), "banos": f.get("banos"),
        "medio_banos": f.get("medios_banos"), "estacionamiento": f.get("estacionamientos"),
        "niveles": f.get("niveles"),
        "antiguedad": (anio_actual - f["anio"]) if f.get("anio") else None,
        "conservacion": f.get("conservacion"), "descripcion": f.get("descripcion"),
        "activo": True, "created_at": ahora, "updated_at": ahora,
    }


def fila_a_doc_pool(f: dict, id_unico: str, *, portal_origen: str, fuente: str,
                     colonia_fuente: str, inmobiliaria_id: str | None, ahora: str,
                     link_verificado: bool | None = None) -> dict:
    """Fila normalizada (con precio+año) → doc de `mercado_props` (pool de comps)."""
    doc = {
        "id_unico": id_unico, "portal_origen": portal_origen, "fuente": fuente,
        "inmobiliaria_id": inmobiliaria_id, "colonia_fuente": colonia_fuente,
        "tipo_propiedad": f["tipo"], "precio": f["precio"],
        "colonia": f["colonia"], "municipio": f["municipio"],
        "anio_construccion": f["anio"], "tipo_operacion": "venta",
        "m2_construccion": f.get("m2_construccion"), "m2_terreno": f.get("m2_terreno"),
        "recamaras": f.get("recamaras"), "banos": f.get("banos"),
        "estacionamientos": f.get("estacionamientos"),
        "activo": True, "fecha_scraping": ahora[:10], "mongo_ts": ahora,
    }
    if f.get("link"):
        doc["url_original"] = f["link"]
        doc["link_verificado"] = link_verificado
    if f.get("grado_remodelacion"):
        doc["grado_remodelacion"] = f["grado_remodelacion"]
        if f.get("anio_remodelacion"):
            doc["anio_remodelacion"] = f["anio_remodelacion"]
    return doc


async def verificar_link(url: str) -> bool | None:
    """Best-effort: intenta confirmar que la URL sigue publicada. NUNCA lanza —
    varios portales (PINCALI) sueltan soft-block a requests automáticos, así que
    un fallo aquí no debe bloquear el guardado, solo se reporta `False`."""
    if not url:
        return None
    import httpx
    try:
        async with httpx.AsyncClient(follow_redirects=True, timeout=6.0) as client:
            r = await client.get(url, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            })
            return r.status_code < 400
    except Exception:
        return False


# ── Descuento por CALIDAD (#142) ─────────────────────────────────────────────
# El descuento lo GANA la calidad, no el volumen crudo: solo cuentan las filas
# NUEVAS (no duplicadas) y completas (la plantilla ya obliga los campos clave).
# Tramos tunables; el descuento es acumulable (se toma el máx histórico).
def descuento_por_calidad(filas_nuevas: int) -> int:
    if filas_nuevas <= 0:
        return 0
    if filas_nuevas < 20:      # ponytail: tramos tunables por negocio
        return 20
    if filas_nuevas < 50:
        return 35
    return 50


# ── Lectura del archivo subido ────────────────────────────────────────────────
def parse_upload(contenido: bytes, filename: str) -> list[dict]:
    """Lee .xlsx o .csv → lista de filas {clave_canónica: valor}. Reconoce las
    columnas por su etiqueta (fila de encabezados); ignora columnas ajenas."""
    nombre = (filename or "").lower()
    if nombre.endswith(".csv"):
        filas_raw = _leer_csv(contenido)
    elif nombre.endswith(".xlsx"):
        filas_raw = _leer_xlsx(contenido)
    else:
        raise ValueError("Formato no soportado: sube .xlsx o .csv")
    if not filas_raw:
        return []
    encabezados = filas_raw[0]
    # índice de columna → clave canónica (por etiqueta, insensible a caso/acentos-espacios)
    col_a_clave = {}
    for i, h in enumerate(encabezados):
        clave = CLAVE_POR_ETIQUETA.get(str(h or "").strip().lower())
        if clave:
            col_a_clave[i] = clave
    if not col_a_clave:
        raise ValueError("No se reconoció ninguna columna de la plantilla. "
                         "Descarga la plantilla y no cambies los encabezados.")
    filas = []
    for row in filas_raw[1:]:
        if not any(str(c or "").strip() for c in row):
            continue  # fila totalmente vacía
        d = {clave: (row[i] if i < len(row) else None) for i, clave in col_a_clave.items()}
        filas.append(d)
    return filas


def _leer_csv(contenido: bytes) -> list[list]:
    texto = contenido.decode("utf-8-sig", errors="replace")
    return [row for row in csv.reader(io.StringIO(texto))]


def _leer_xlsx(contenido: bytes) -> list[list]:
    import openpyxl  # ponytail: import perezoso, solo si suben xlsx
    wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    ws = wb.active
    filas = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    # Saltar filas de instrucciones anteriores al encabezado real: buscar la fila
    # que contenga "Tipo de propiedad".
    for idx, row in enumerate(filas):
        if any(str(c or "").strip().lower() == "tipo de propiedad" for c in row):
            return filas[idx:]
    return filas


# ── Generación de la plantilla en blanco ──────────────────────────────────────
def generar_plantilla_xlsx() -> bytes:
    """Plantilla .xlsx: fila de instrucciones, encabezados y una fila de ejemplo."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    etiquetas = [ETIQUETA[k] for k, _ in COLUMNAS]
    ws.append(["INSTRUCCIONES: llena una fila por propiedad. Para Terreno solo se "
               "piden Tipo, Dirección, Colonia, Municipio, Precio y m² terreno. "
               "No cambies los encabezados de abajo."] + [""] * (len(etiquetas) - 1))
    ws.append(etiquetas)
    ws.append([EJEMPLO.get(k, "") for k, _ in COLUMNAS])
    encabezado_fill = PatternFill("solid", fgColor="1B4332")
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = encabezado_fill
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
